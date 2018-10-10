'''
    File name: app.py
    Author: Batjargal (Alex) Bayarsaikhan
    Date created: 09/26/2018
    Date last modified: 10/02/2018
    Python Version: 3.6
'''

import getpass
import time
import requests
import json
import mechanicalsoup
import numpy as np

import subprocess

import os
import win32com.client 
import openpyxl

from datetime import datetime, timedelta

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def copy2clip(txt):
    cmd='echo '+txt.strip()+'|clip'
    return subprocess.check_call(cmd, shell=True)

	
def simple_decode(vl):
	result = ""
	for ch in vl:
		result += chr(ord(ch) + 15)
	return result
	
def is_valid_date(date_str):
    try:
        datetime.strptime(date_str, '%m/%d/%Y')
        return True
    except ValueError:
        print("Incorrect date format, should be mm/dd/yyyy")
    return False

def get_date():
	current_date = input("Please enter date (mm/dd/yyyy): ")
	if not is_valid_date(current_date):
		return get_date()
	return current_date	

def pull_data(config, current_date):
	
	services = {}
	
	address = config["address"]
	print(f"Connect to {address}")
	
	username = input("Please enter your username: ")
	password = getpass.getpass("Please enter your password: ")
	#username = "__USER_NAME__" #

	print(f"Login to {address}")
	browser = mechanicalsoup.StatefulBrowser()
	browser.open(address + "/__HOME_PAGE_URL_IF_REQUIRED_")

	browser.select_form('form[name="__LOGIN_FORM_NAME__"]')
	browser["username"] = username
	browser["password"] = password
	browser.submit_selected()
	
	
	try:
		browser.select_form('form[name="__LOGIN_FORM_NAME__"]')
		print("Username or password is incorrect.")			
		return pull_data(config, current_date)
	except:
		pass
	
	followlink = config["followlink"]
			
	groups = config["groups"]
	
	
	print("Pulling data from API services...")
	
	for group in groups:
		group_name = group["name"]
		group_suffix = group["suffix"]
		sources = group["sources"]
		for source in sources:
			print(source["uddi"])
			service_name = source["name"].replace(group_suffix, "").strip().lower()
			services[service_name] = {}
			browser.open(f'{address}/__URL_PATH__?objId=uddi:{source["uddi"]}&csrf-token=0xa9999999')

			
			browser.select_form('#__FORM_NAME__')
			browser["__FORM_NAME__:__VALUE_NAME__1"] = "__VALUE_DATA__1"
			browser["__FORM_NAME__:__VALUE_NAME__2"] = "__VALUE_DATA__2"
			browser["__FORM_NAME__:__VALUE_NAME__3"] = current_date
			browser["__FORM_NAME__:__VALUE_NAME__4"] = "__VALUE_DATA__4"
			browser.submit_selected()
			
			# only for debugging purpose
			#if source["uddi"] == "__UNIQUE_ID__":
			#	print(service_name + "              ===========================")
			#	browser.launch_browser()
				
			page = browser.get_current_page()
			table = page.find("table", class_="tableLists")
			rows = table.find_all("tr")[1:]
			for row in rows:
				tds = row.find_all("td")
				if tds[1].text.strip().lower() != "not found":
					services[service_name][tds[1].text.strip().lower()] = [tds[2].text, tds[3].text]
				
	return services
	
def working_with_excel(services, file_name, current_date):
						 
	current_service_name = ""
	current_operation = ""
	
	thin_border = Border(left=Side(style='thin'), 
						 right=Side(style='thin'), 
						 top=Side(style='thin'), 
						 bottom=Side(style='thin'))
						 
	left_alignment = Alignment(horizontal='left',
						 vertical='center',
						 text_rotation=0,
						 wrap_text=False,
						 shrink_to_fit=False,
						 indent=0)
						 
	center_top_alignment = Alignment(horizontal='center',
						 vertical='top',
						 text_rotation=0,
						 wrap_text=False,
						 shrink_to_fit=False,
						 indent=0)
						 
	wb = openpyxl.load_workbook(filename=file_name, read_only=False, keep_vba=True)

	sheet = wb["Sheet1"]

	new_column_index = int(sheet.max_column) - 2
	sheet.insert_cols(new_column_index)
	
	# hide column
	#sheet.column_dimensions[get_column_letter(new_column_index - 8)].hidden= True
	

	new_cell = sheet.cell(row=3, column=new_column_index)
	new_cell.value = current_date
	new_cell.fill = PatternFill(fgColor='305496', fill_type = 'solid')
	new_cell.font=Font(color='FFFFFF', size=12, bold=True)
	new_cell.border = thin_border
	new_cell.alignment=Alignment(horizontal='general',
						 vertical='top',
						 text_rotation=0,
						 wrap_text=False,
						 shrink_to_fit=False,
						 indent=0)


	sheet.column_dimensions[get_column_letter(new_column_index)].width = 11.86
		
	col_ind_known = new_column_index + 1	# Known Errors 
	col_ind_status = new_column_index + 2	# Overall Status
	col_ind_comment = new_column_index + 3	# Comment	
	
	sheet.column_dimensions[get_column_letter(col_ind_known)].width = 50.00	
	sheet.column_dimensions[get_column_letter(col_ind_status)].width = 35.14	
	sheet.column_dimensions[get_column_letter(col_ind_comment)].width = 30.14	

	for row_index in range(4, sheet.max_row):		
		
		sheet.cell(row=row_index, column=1).border = thin_border	# Service
		sheet.cell(row=row_index, column=2).border = thin_border	# Operation
		
		sheet.cell(row=row_index, column=col_ind_known).border = thin_border
		sheet.cell(row=row_index, column=col_ind_comment).border = thin_border
		
		
		cell_overal_status = sheet.cell(row=row_index, column=col_ind_status)
		cell_overal_status.border = thin_border
		cell_overal_status.alignment = center_top_alignment
		cell_overal_status.value = "Good"		
		cell_overal_status.fill = PatternFill(fgColor='c6efce', fill_type = 'solid')
		
		operation_found = False
		service_name = sheet.cell(row=row_index, column=1).value
		
		if service_name:
			service_name = service_name.strip().lower()
			current_service_name = service_name
		
		if current_service_name not in services:
			current_service_name = "_".join(current_service_name.split('_')[:-1])
			
		if current_service_name in services:
					
			operation = sheet.cell(row=row_index, column=2).value
			if operation:
				operation = operation.strip().lower()
				if operation in services[current_service_name]:
					operation_found = True
					current_operation = services[current_service_name][operation]
					
					sheet.unmerge_cells(start_row=row_index, start_column=new_column_index, end_row=row_index + 3, end_column=new_column_index)
					sheet.merge_cells(start_row=row_index, start_column=new_column_index + 3, end_row=row_index + 3, end_column=new_column_index + 3)
					
					cell_failed = sheet.cell(row=row_index, column=new_column_index)
					cell_failed.value = current_operation[1]
					cell_failed.border = thin_border
					
					# error number
					if int(current_operation[1]) >= 100:					
						cell_overal_status.value = "More Deviation(Risk)"		
						cell_overal_status.fill = PatternFill(fgColor='FFFF00', fill_type = 'solid')
					elif int(current_operation[1]) >= 11:
						cell_overal_status.value = "Less Deviation(Expected)"	
						cell_overal_status.fill = PatternFill(fgColor='ffcc99', fill_type = 'solid')
						
					
					cell_success = sheet.cell(row=row_index + 1, column=new_column_index)
					cell_success.value = current_operation[0]
					cell_success.border = thin_border
					
					cell_total = sheet.cell(row=row_index + 2, column=new_column_index)
					cell_total.value = int(current_operation[0]) + int(current_operation[1])
					cell_total.border = thin_border					
					cell_total.alignment = left_alignment						 
						 
					cell_succ_percent = sheet.cell(row=row_index + 3, column=new_column_index)					
					cell_succ_percent.border = thin_border
					cell_succ_percent.number_format = '0.00%'
					cell_succ_percent.alignment = left_alignment
					cell_succ_percent.value = 0
					
					#print(current_service_name)
				else:
					sheet.unmerge_cells(start_row=row_index, start_column=new_column_index, end_row=row_index + 3, end_column=new_column_index)
					sheet.merge_cells(start_row=row_index, start_column=new_column_index + 3, end_row=row_index + 3, end_column=new_column_index + 3)
					
			if not operation_found:
				cell_empty = sheet.cell(row=row_index, column=new_column_index)
				# skip cells which have vale
				if cell_empty.value == "" or cell_empty.value is None:		
					cell_empty.value = "0"
					#cell_empty.fill = PatternFill(fgColor='FFFF00', fill_type = 'solid')
					cell_empty.border = thin_border
				
					cell_succ_percent = sheet.cell(row=row_index, column=3)					
					if cell_succ_percent.value.strip().lower() == "success %":
						cell_empty.number_format = '0.00%'
						cell_empty.alignment = left_alignment
						
			
					
		else:	
			cell_empty = sheet.cell(row=row_index, column=new_column_index)		
			cell_empty.fill = PatternFill(fgColor='FF8899', fill_type = 'solid')
				
			

	wb.save(file_name)



def run_macro(file_name):
	dir_path = os.path.dirname(os.path.realpath(__file__))
	if os.path.exists(file_name):
		xl = win32com.client.Dispatch('Excel.Application')
		wb = xl.Workbooks.Open(Filename = f"{dir_path}\\{file_name}", ReadOnly=1)
		xl.Application.Run("__MACRO_NAME__")
		wb.Save()
		xl.Application.Quit()
		xl.Quit()
		del xl

#import numpy as np

def main():

	try:
		print("--- The tool: Version 0.1, 10/02/2018 ---")
		
		config = {}
		with open("url_info.json", "r") as f:
			config = json.load(f)

		current_date = get_date()
		services = pull_data(config, current_date)
			
		file_name = '__FILE_NAME__'

		print("Working with excel...")
		working_with_excel(services, file_name, current_date)
		
		#services = np.load('services_temp.npy')
		#working_with_excel(services.item(), file_name, current_date)
		
		print("Running Macro...")
		run_macro(file_name)
		
		print("All operations are completed.")
		
	except Exception as ex:
		print(ex)
		print("Please try again.")
	


password = ""

#with open("password", "r") as f:
#	password = simple_decode(f.readline().strip())

#copy2clip(password)
#browser.launch_browser()
#print(browser.get_current_page())
#print(browser.get_url())
#print(messages)
	
#yesterday = datetime.now() - timedelta(days=1)
#current_date = f"{yesterday.month}/{yesterday.day}/{yesterday.year}"

#services = pull_data(config, current_date)
#np.save('services_temp.npy', services) 

#file_name = '__FILE_NAME__'
#services = np.load('services_temp.npy')	

if __name__ == "__main__":
	main()
